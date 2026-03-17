"""
Analytics Reader
----------------
Reads the weekly LinkedIn Analytics Excel export from data/analytics/
and returns structured performance data for the assessment agent.

Expected export format (LinkedIn Creator Analytics download):
  Sheet: DISCOVERY   — overall weekly impressions + members reached
  Sheet: ENGAGEMENT  — daily impressions + engagements
  Sheet: TOP POSTS   — per-post impressions and engagements (up to 50 posts)
  Sheet: FOLLOWERS   — total followers + daily new followers
  Sheet: DEMOGRAPHICS — audience breakdown by job title etc.

Usage:
    from agents.analytics_reader import load_latest_analytics, format_for_assessment
    data = load_latest_analytics()          # None if no file found
    text = format_for_assessment(data, article_date="2026-03-10")
"""

import re
from datetime import datetime, date
from pathlib import Path
from typing import Optional

ANALYTICS_DIR = Path(__file__).parent.parent / "data" / "analytics"


# ---------------------------------------------------------------------------
# Excel parsing
# ---------------------------------------------------------------------------

def _parse_date(value) -> Optional[date]:
    """Parse M/D/YYYY or M/D/YY strings (as returned by openpyxl) into date."""
    if not value:
        return None
    s = str(value).strip()
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _parse_top_posts(ws) -> list[dict]:
    """
    TOP POSTS sheet has two side-by-side tables:
      Left  (cols 0,1,2): Post URL | Post publish date | Engagements
      Right (cols 4,5,6): Post URL | Post publish date | Impressions

    Merge both by URL → {url, date, engagements, impressions}
    """
    by_url: dict[str, dict] = {}

    for row in ws.iter_rows(min_row=4, values_only=True):  # skip header rows
        # Left side (engagement-ranked)
        url_l, date_l, eng = row[0], row[1], row[2]
        if url_l and str(url_l).startswith("http"):
            entry = by_url.setdefault(url_l, {"url": url_l})
            entry["date"]        = _parse_date(date_l)
            entry["engagements"] = int(eng) if eng else 0

        # Right side (impression-ranked)
        url_r, date_r, imp = row[4], row[5], row[6]
        if url_r and str(url_r).startswith("http"):
            entry = by_url.setdefault(url_r, {"url": url_r})
            entry.setdefault("date", _parse_date(date_r))
            entry["impressions"] = int(imp) if imp else 0

    # Fill missing fields
    posts = []
    for p in by_url.values():
        p.setdefault("engagements", 0)
        p.setdefault("impressions", 0)
        if p.get("date"):
            posts.append(p)

    return sorted(posts, key=lambda x: x.get("impressions", 0), reverse=True)


def _parse_discovery(ws) -> dict:
    """Returns {impressions, members_reached, period}."""
    result = {}
    period_raw = None
    for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
        if row[0] == "Overall Performance" and row[1]:
            period_raw = str(row[1])
        if row[0] == "Impressions" and row[1]:
            result["impressions"] = int(row[1])
        if row[0] == "Members reached" and row[1]:
            result["members_reached"] = int(row[1])
    if period_raw:
        result["period"] = period_raw
    return result


def _parse_engagement(ws) -> list[dict]:
    """Returns [{date, impressions, engagements}, ...]."""
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = _parse_date(row[0])
        if d and row[1] is not None:
            rows.append({
                "date":        d,
                "impressions": int(row[1]),
                "engagements": int(row[2]) if row[2] else 0,
            })
    return rows


def _parse_followers(ws) -> dict:
    """Returns {total, new_this_week}."""
    total = None
    new_followers = 0
    data_started = False
    for row in ws.iter_rows(values_only=True):
        if row[0] and "Total followers" in str(row[0]):
            try:
                total = int(row[1])
            except (TypeError, ValueError):
                pass
        if row[0] == "Date":
            data_started = True
            continue
        if data_started and row[1] is not None:
            try:
                new_followers += int(row[1])
            except (TypeError, ValueError):
                pass
    return {"total": total, "new_this_week": new_followers}


# ---------------------------------------------------------------------------
# Public loader
# ---------------------------------------------------------------------------

def load_latest_analytics() -> Optional[dict]:
    """
    Find the most recently modified .xlsx file in data/analytics/ and parse it.
    Returns None if no file exists or openpyxl is not installed.
    """
    files = sorted(ANALYTICS_DIR.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    if not files:
        return None

    try:
        import openpyxl
    except ImportError:
        print("  ⚠️  openpyxl not installed — run: pip3 install openpyxl")
        return None

    path = files[0]
    wb = openpyxl.load_workbook(path, data_only=True)

    data = {
        "file":        path.name,
        "posts":       [],
        "discovery":   {},
        "engagement":  [],
        "followers":   {},
    }

    if "TOP POSTS"  in wb.sheetnames: data["posts"]      = _parse_top_posts(wb["TOP POSTS"])
    if "DISCOVERY"  in wb.sheetnames: data["discovery"]  = _parse_discovery(wb["DISCOVERY"])
    if "ENGAGEMENT" in wb.sheetnames: data["engagement"] = _parse_engagement(wb["ENGAGEMENT"])
    if "FOLLOWERS"  in wb.sheetnames: data["followers"]  = _parse_followers(wb["FOLLOWERS"])

    return data


# ---------------------------------------------------------------------------
# Format for assessment prompt
# ---------------------------------------------------------------------------

def format_for_assessment(data: Optional[dict], article_date: Optional[str] = None) -> str:
    """
    Render analytics data as a markdown block for inclusion in the assessment prompt.

    article_date: ISO format "YYYY-MM-DD" — used to look up this specific post's metrics.
    Returns empty string if no data available.
    """
    if not data:
        return ""

    lines = [f"## LinkedIn Analytics  ({data.get('file', 'export')})"]

    # --- Weekly overview ---
    disc = data.get("discovery", {})
    if disc:
        lines.append(f"\n### Week overview  ({disc.get('period', '')})")
        lines.append(f"- Total impressions: **{disc.get('impressions', '?')}**")
        lines.append(f"- Members reached: **{disc.get('members_reached', '?')}**")

    fol = data.get("followers", {})
    if fol.get("total"):
        lines.append(f"- Followers: **{fol['total']:,}** (+{fol.get('new_this_week', 0)} this week)")

    # --- This post's metrics ---
    target_date = None
    if article_date:
        try:
            target_date = datetime.strptime(article_date, "%Y-%m-%d").date()
        except ValueError:
            pass

    matched = None
    if target_date:
        for p in data.get("posts", []):
            if p.get("date") == target_date:
                matched = p
                break

    if matched:
        total_imp = disc.get("impressions", 0)
        imp = matched.get("impressions", 0)
        eng = matched.get("engagements", 0)
        eng_rate = f"{eng/imp*100:.1f}%" if imp else "n/a"
        share    = f"{imp/total_imp*100:.0f}%" if total_imp else "n/a"

        # Rank among all posts this week
        sorted_by_imp = sorted(data["posts"], key=lambda x: x.get("impressions", 0), reverse=True)
        rank_imp = next((i+1 for i, p in enumerate(sorted_by_imp) if p["url"] == matched["url"]), "?")
        sorted_by_eng = sorted(data["posts"], key=lambda x: x.get("engagements", 0), reverse=True)
        rank_eng = next((i+1 for i, p in enumerate(sorted_by_eng) if p["url"] == matched["url"]), "?")

        lines.append(f"\n### This post's performance  (published {matched['date']})")
        lines.append(f"- Impressions: **{imp}** (#{rank_imp} of {len(data['posts'])} tracked posts, {share} of weekly total)")
        lines.append(f"- Engagements: **{eng}** (#{rank_eng} of {len(data['posts'])} tracked posts)")
        lines.append(f"- Engagement rate: **{eng_rate}**")
        lines.append(f"- Post URL: {matched['url']}")
    elif target_date:
        lines.append(f"\n_No analytics match found for {target_date} — post may not appear in the top-50 list or was published outside the export window._")

    # --- Top 3 posts for context ---
    posts = data.get("posts", [])
    if posts:
        lines.append("\n### Top posts this week (by impressions)")
        for i, p in enumerate(posts[:3], 1):
            marker = " ← this post" if p is matched else ""
            lines.append(f"{i}. {p['date']} — {p.get('impressions',0)} impr / {p.get('engagements',0)} eng{marker}")

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Standalone test
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    data = load_latest_analytics()
    if data:
        print(f"Loaded: {data['file']}")
        print(f"Posts found: {len(data['posts'])}")
        print()
        print(format_for_assessment(data, article_date="2026-03-10"))
    else:
        print("No analytics file found in data/analytics/")
